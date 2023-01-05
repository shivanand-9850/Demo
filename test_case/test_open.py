import openpyxl
# ws=openpyxl.load_workbook('C:/Users/vitth/OneDrive/Desktop/shiva/Automation_project_1/practice_Excel.xlsx')
# sht=ws.active
# c=sht.cell(row=3,column=2)
# print(c.value)
from selenium import webdriver
import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common import action_chains as act
driver=webdriver.Chrome()


@pytest.fixture()
def setup():
    global driver
    driver=webdriver.Chrome()
    driver.maximize_window()
    yield
    print(driver.title)
    driver.close()
    # print('Get driver')
    # print('Maximize window')
    # yield
    # print('print title')
    # print('close window')

def test_1(setup):
    driver.get('https://www.facebook.com/')
    print("poen Url of facebook")

def test_2(setup):
    driver.get('https://www.google.com/')
    print("poen Url of gmail"'ht://www.google.com/')

def test_3(setup):
    driver.get('https://www.amazon.com/')
    print("poen Url of amazon")


def test_4(setup):
    driver.get('https://www.jio.com/')

filname='C:/Users/vitth/OneDrive/Desktop/shiva/Automation_project_1/practice_Excel.xlsx'
workbook=openpyxl.load_workbook(filname)
sheet=workbook['sheet1']
print(sheet.cell(row=4,column=2).value)
print(sheet.cell(row=5,column=2).value)
temp =sheet.cell(row=1, column=2).value
sheet.cell(row=6,column=1).value ='Sitaram'
sheet.cell(row=6,column=2).value ='shiv'
workbook.save(filname)
