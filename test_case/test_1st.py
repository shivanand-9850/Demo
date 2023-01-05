import time
import openpyxl
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
# from selenium.webdriver.common.
from selenium.webdriver.common.action_chains import ActionChains as action

driver = webdriver.Chrome()
# driver.get('https://www.google.com/')
# act = action(driver)
# a = driver.find_element(By.XPATH, '//img[@alt="Google"]')
# act.double_click(a).perform()
# act.context_click(a).perform()

# act.send_keys(Keys.ARROW_DOWN)
# act.send_keys(Keys.ARROW_DOWN).perform()
# time.sleep(5)
# w = driver.window_handles
# print(w)
# import pyautogui
#
# pyautogui.press('down')
# pyautogui.press('enter')

ws = openpyxl.load_workbook(r'C:/Users/vitth/OneDrive/Desktop/shiva/Automation_project_1/practice_Excel.xlsx')
sht = ws.active
c = sht.cell(row=3, column=3)
print(sht.max_row)
print(c.value)
